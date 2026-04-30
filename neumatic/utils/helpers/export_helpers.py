# neumatic\utils\helpers\export_helpers.py
import os
from dotenv import load_dotenv
from io import BytesIO, TextIOWrapper
from reportlab.platypus import BaseDocTemplate, SimpleDocTemplate, Frame, PageTemplate, LongTable, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4, portrait
from reportlab.lib.enums import TA_LEFT, TA_RIGHT
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib import colors
from openpyxl import Workbook, styles
import csv
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from datetime import datetime
from decimal import Decimal
from functools import partial
from django.core.exceptions import PermissionDenied

from apps.maestros.templatetags.custom_tags import formato_es_ar
from utils.utils import format_date, formato_argentino, formato_argentino_entero
from entorno.constantes_base import JERARQUIAS_CON_ACCESO_TOTAL

#-- Cargar las variables de entorno del archivo .env
load_dotenv()


class ExportHelper:
	
	def __init__(self, queryset, table_info, report_title, total_columns=None):
		self.queryset = queryset
		self.table_info = table_info
		self.report_title = report_title
		self.total_columns = total_columns if total_columns else {}  #-- Diccionario con texto y columnas a totalizar.
	
	def _safe_str(self, value):
		return str(value) if value is not None else ""
	
	def _get_headers_and_fields(self):
		"""Obtener nombres de encabezados y los campos correspondientes del diccionario table_info."""
		
		headers = [field['label'] for field in self.table_info.values()]
		# fields = [ field for field in self.table_info.keys()]   #-- Menos eficiente que el siguiente método.
		fields = list(self.table_info.keys())   #-- Más eficiente por ser más directo.
		protected = [field for field in self.table_info if self.table_info[field].get('protected', False)]
		
		return headers, fields, protected
	
	def _resolve_field(self, obj, field_name, frmto=None):
		"""
		Resuelve el valor de un campo (incluso anidado).
		Si frmto es 'pdf', se formatea el valor (por ejemplo, números se convierten a cadena formateada);
		en caso contrario, se devuelve el valor en su tipo original (para Excel/CSV) para que los datos numéricos
		sean numéricos (enteros y Decimal con 2 decimales).
		"""
		try:
			#-- Recorrer los niveles anidados para obtener el valor.
			fields = field_name.split('.')
			value = obj
			for f in fields:
				#-- Si el atributo existe, se obtiene.
				if hasattr(value, f):
					value = getattr(value, f)
				#-- Sino, se intenta obtener del diccionario interno (__dict__).
				elif isinstance(value, dict) and f in value:
					value = value[f]
				else:
					value = None
					break
					
			if value is None:
				return ""
					
			#-- Si es una instancia de un modelo.
			if hasattr(value, '_meta') and frmto == 'excel':
				return str(value)  # Usa la representación en string del objeto (definida en __str__).			
			
			#-- Si es booleano.
			if isinstance(value, bool):
				if frmto == 'pdf':
					#-- Para PDF, se formatea según si el campo contiene "estatus".
					if "estatus" in fields[-1]:
						return "Activo" if value else "Inactivo"
					else:
						return "Sí" if value else "No"
				else:
					return value
			
			#-- Si es numérico.
			if isinstance(value, (float, Decimal)):
				if frmto == 'pdf':
					#-- Para PDF, formatea el número a cadena.
					return self._safe_str(formato_es_ar(value))
				else:
					#-- Para Excel/CSV, devuelve el número:
					#-- - Si es float, redondea a 2 decimales.
					#-- - Si es Decimal, lo cuantiza a 2 decimales.
					if isinstance(value, float):
						return round(value, 2)
					else:
						return value.quantize(Decimal('0.01'))
			
			#-- Otros tipos.
			if frmto == 'pdf':
				return self._safe_str(value)
			else:
				return value
		except Exception as e:
			return ""
	
	def _calculate_totals(self, fields):
		"""Calcula los totales para las columnas especificadas en total_columns y genera la fila de totales."""
		
		totals = [""] * len(fields)  #-- Inicializa la fila de totales vacía.
		
		for total_text, columns in self.total_columns.items():
			#-- Encuentra la primera columna para colocar el texto.
			first_col = None
			for col in columns:
				if col in fields:
					col_index = fields.index(col)
					total = sum(
						getattr(obj, col.split('.')[0], 0) or 0
						for obj in self.queryset
					)
					#-- Formatear el total.
					totals[col_index] = formato_es_ar(total)
					
					if first_col is None:  # Coloca el texto solo una vez
						first_col = col_index
			
			#-- Coloca el texto en la primera columna encontrada.
			if first_col is not None:
				totals[first_col - 1 if first_col > 0 else 0] = total_text
		
		return totals
	
	def export_to_pdf(self, pagesize=portrait(A4), margins=(20, 20, 20, 40), body_font_size=8):
		"""Exporta los datos del queryset a un archivo PDF."""
		
		left_margin, right_margin, top_margin, bottom_margin = margins
		buffer = BytesIO()
		
		#-- Estilo de párrafo para los datos en la tabla del reporte.
		styles = getSampleStyleSheet()
		styles.add(ParagraphStyle(
			name='CellStyle',
			parent=styles["BodyText"],
			fontSize=body_font_size,
			leading=8,
			spaceBefore=0,
			spaceAfter=0,
			leftIndent=0,
			rightIndent=0,
			firstLineIndent=0,
		))
		
		paragraph_style = styles['CellStyle']
		
		doc = SimpleDocTemplate(
			buffer,
			pagesize=pagesize,
			leftMargin=left_margin,
			rightMargin=right_margin,
			topMargin=top_margin,
			bottomMargin=bottom_margin
		)
		elements = []
		
		#-- Encabezado del reporte.
		elements.append(Paragraph(self.report_title, getSampleStyleSheet()['Title']))
		elements.append(Spacer(1, 12))
		
		#-- Obtener encabezados y campos dinámicos.
		headers, fields, protected = self._get_headers_and_fields()
		
		cols_widths = [field['col_width_pdf'] for field in self.table_info.values()]
		
		paragraph_fields = [field for field in self.table_info.keys() if self.table_info[field]['pdf_paragraph']]
		
		#-- Crear la tabla de datos.
		table_data = [headers]  #-- Primera fila con los encabezados.
		
		#-- Agregar los datos desde el queryset.
		for obj in self.queryset:
			row = []
			for field in fields:
				value = self._resolve_field(obj, field, frmto='pdf')
				if field in paragraph_fields and value:
					row.append(Paragraph(value, paragraph_style))
				else:
					if self.table_info[field]['date_format']:
						row.append(format_date(value))
					else:
						row.append(value)
			table_data.append(row)
		
		#-- Calcular y agregar la fila de totales.
		if self.total_columns:
			totals_row = self._calculate_totals(fields)
			table_data.append(totals_row)
		
		table = LongTable(table_data, colWidths=cols_widths)
		
		table_style = TableStyle([
			#-- Estilos comunes toda la tabla.
			('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
			('FONTSIZE', (0,0), (-1,-1), body_font_size),
			('LEADING', (0,0), (-1,-1), 8),
			('VALIGN', (0,0), (-1,-1), 'TOP'),
			
			#-- Padding de la tabla exceptuando la primera fila (headers).
			('TOPPADDING', (0,1), (-1,-1), 0),
			('BOTTOMPADDING', (0,1), (-1,-1), 0),
			
			#-- Estilos para la primera fila (headers).
			('BACKGROUND', (0,0), (-1,0), colors.gray),
			('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
			('TEXTCOLOR', (0,0), (-1,0), colors.white),
			('TOPPADDING', (0,0), (-1,0), 2),
			('BOTTOMPADDING', (0,0), (-1,0), 2),
			
			# Línea horizontal por encima de los títulos
			# ('LINEABOVE', (0,0), (-1,0), 0.5, colors.black),
			# Línea horizontal por debajo de los títulos
			# ('LINEBELOW', (0,0), (-1,0), 0.5, colors.black),
			# ('TEXTCOLOR', (0,0), (-1,0), colors.black),
		])
			
		#-- Identificar las columnas numericas y/o booleanas para alinear a la derecha o al centro el dato y el header.
		cols_headers_right = []
		cols_headers_center = []
		for col_idx, field in enumerate(fields):
			if any(isinstance(getattr(obj, field.split('.')[0], None), (int, float, Decimal)) for obj in self.queryset):
				table_style.add('ALIGN', (col_idx, 1), (col_idx, -1), 'RIGHT')
				cols_headers_right.append(col_idx)
			if isinstance(self._resolve_field(obj, field, frmto='excel'), bool):
				table_style.add('ALIGN', (col_idx, 0), (col_idx, -1), 'CENTER')
				cols_headers_center.append(col_idx)
		
		#-- Alinear a la derecha los headers de las columnas numéricas.
		for col in cols_headers_right:
			table_style.add('ALIGN', (col,0), (col,0), 'RIGHT')
		
		#-- Alinear al contro los headers de las columnas booleanas.
		for col in cols_headers_center:
			table_style.add('ALIGN', (col,0), (col,0), 'CENTER')
		
		#-- Alinear la fila de totales y aplicar negritas.
		if self.total_columns:
			totals_row_index = len(table_data) - 1  # Índice de la fila de totales (última fila)
			
			#-- Negritas en toda la fila de totales.
			table_style.add('FONTNAME', (0, totals_row_index), (-1, totals_row_index), 'Helvetica-Bold')
			
			#-- Agregar línea horizontal encima de la fila de totales.
			table_style.add('LINEABOVE', (0, -1), (-1, -1), 1, colors.black)
			
			#-- Alinear columnas numéricas a la derecha en la fila de totales.
			for total_text, columns in self.total_columns.items():
				for col in columns:
					if col in fields:
						col_index = fields.index(col)
						table_style.add('ALIGN', (col_index, -1), (col_index, -1), 'RIGHT')
			
		table.setStyle(table_style)
		
		#-- Repite la primera fila (headers) en cada página.
		table.repeatRows = 1
		
		elements.append(table)
		
		#-- Pasar los márgenes a CustomCanvas usando partial.
		canvas_with_margins = partial(
			CustomCanvas, 
			margins=(left_margin, right_margin, top_margin, bottom_margin)
		)
		
		doc.build(elements, canvasmaker=canvas_with_margins)
		buffer.seek(0)
		
		return buffer.getvalue()
	
	def export_to_excel(self):
		"""Exporta los datos del queryset a un archivo Excel."""
		
		wb = Workbook()
		ws = wb.active
		ws.title = self.report_title[:31]  #-- Limitar a 31 caracteres (restricción de Excel).
		
		#-- Obtener encabezados y campos dinámicos.
		headers, fields, protected = self._get_headers_and_fields()
		
		#-- Encabezados.
		ws.append(headers)
		
		#-- Identificar índices de campos a proteger
		protected_indices = []
		if protected:
			protected_indices = [fields.index(field) for field in protected if field in fields]
		
		#-- Procesar datos según el tipo de queryset.
		if isinstance(self.queryset, dict):
			#-- Caso cuando el queryset es un diccionario.
			for key, data in self.queryset.items():
				if isinstance(data, dict):
					#-- Para diccionarios anidados.
					row = []
					for field in fields:
						#-- Manejar campos anidados.
						value = data
						for part in field.split('.'):
							if isinstance(value, dict) and part in value:
								value = value[part]
							elif hasattr(value, part):
								value = getattr(value, part)
							else:
								value = None
								break
						row.append(self._resolve_field(value, '', frmto='excel'))
					ws.append(row)
				else:
					#-- Para otros tipos de valores en el diccionario.
					row = [self._resolve_field(data, field, frmto='excel') for field in fields]
					ws.append(row)
		else:
			#-- Caso normal (RawQuerySet, QuerySet, lista).
			for obj in self.queryset:
				row = [self._resolve_field(obj, field, frmto='excel') for field in fields]
				ws.append(row)
		
		#-- Proteger campos específicos si hay campos a proteger.
		if protected_indices:
			#-- Bloquear los títulos de columnas (fila 1 - encabezados).
			for col_idx, cell in enumerate(ws[1], 1):  #-- ws[1] es la primera fila (encabezados).
				cell.protection = styles.Protection(locked=True)
			
			#-- Bloquear solo las celdas específicas que necesitan protección
			for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):  #-- Empezar desde la fila 2 (datos) omitiendo encabezados.
				for col_idx, cell in enumerate(row, 1):
					if (col_idx - 1) in protected_indices:  #-- -1 porque las columnas empiezan en 1 pero los índices en 0.
						cell.protection = styles.Protection(locked=True)
					else:
						cell.protection = styles.Protection(locked=False)
			
			#-- Proteger la hoja.
			ws.protection.enable()
			ws.protection.sheet = True
			
			#-- PERMISOS ESPECÍFICOS - Permitir estas acciones (False = permitido, True = no permitido).
			ws.protection.formatCells = False          # Permitir formato de celdas.
			ws.protection.formatColumns = False        # Permitir formato de columnas.
			ws.protection.formatRows = False           # Permitir formato de filas.
			ws.protection.sort = False                 # Permitir ordenar.
			
			# #-- Restringir otras acciones (opcional, según necesites).
			# ws.protection.selectLockedCells = True     # No permitir seleccionar celdas bloqueadas.
			# ws.protection.selectUnlockedCells = False  # Permitir seleccionar celdas desbloqueadas.
			# ws.protection.insertRows = False           # No permitir insertar filas.
			# ws.protection.insertColumns = False        # No permitir insertar columnas.
			# ws.protection.deleteRows = False           # No permitir eliminar filas.
			# ws.protection.deleteColumns = False        # No permitir eliminar columnas.
			# ws.protection.objects = False              # No permitir editar objetos.
			# ws.protection.scenarios = False            # No permitir escenarios.
			# ws.protection.autoFilter = False           # No permitir filtros automáticos.
			# ws.protection.pivotTables = False          # No permitir tablas dinámicas.
			
			if os.getenv('EXCEL_KEY'):
				ws.protection.password = os.getenv('EXCEL_KEY')
		
		buffer = BytesIO()
		wb.save(buffer)
		buffer.seek(0)
		
		return buffer.getvalue()
	
	def export_to_csv(self):
		"""Exporta los datos del queryset a un archivo CSV, compatible con RawQuerySet y diccionarios."""
		
		#-- Crear buffer binario.
		buffer = BytesIO()
		
		#-- Crear envoltura de texto.
		text_buffer = TextIOWrapper(buffer, encoding="utf-8", newline="")
		
		#-- Escribir el BOM manualmente para indicar UTF-8.
		text_buffer.write('\ufeff')  #-- Agregar el BOM al archivo como texto.
		
		writer = csv.writer(text_buffer)
		
		#-- Obtener encabezados y campos.
		headers, fields, protected = self._get_headers_and_fields()
		
		#-- Encabezados.
		writer.writerow(headers)  #-- Escribir encabezados.
		
		#-- Procesar datos según el tipo de queryset.
		if isinstance(self.queryset, dict):
			#-- Caso cuando el queryset es un diccionario.
			for key, data in self.queryset.items():
				if isinstance(data, dict):
					#-- Para diccionarios anidados.
					row = []
					for field in fields:
						value = data
						for part in field.split('.'):
							if isinstance(value, dict) and part in value:
								value = value[part]
							elif hasattr(value, part):
								value = getattr(value, part)
							else:
								value = None
								break
						row.append(self._resolve_field(value, ''))  #-- Sin formato especial para CSV.
					writer.writerow(row)
				else:
					#-- Para valores directos.
					row = [self._resolve_field(data, field) for field in fields]
					writer.writerow(row)
		else:
			#-- Caso tradicional (RawQuerySet, QuerySet, lista).
			for obj in self.queryset:
				row = [self._resolve_field(obj, field) for field in fields]
				writer.writerow(row)
		
		#-- Vaciar contenido al buffer binario.
		text_buffer.flush()
		buffer.seek(0)
		
		#-- Obtener el contenido en bytes.
		csv_bytes = buffer.getvalue()		
		
		#-- Cerrar buffers.
		text_buffer.close()
		buffer.close()
		
		return csv_bytes


class CustomCanvas(Canvas):
	# def __init__(self, *args, margins=(20, 20, 20, 40), **kwargs):
	def __init__(self, *args, margins, **kwargs):
		super().__init__(*args, **kwargs)
		self.left_margin, self.right_margin, self.top_margin, self.bottom_margin = margins
		self.report_date = datetime.now().strftime("%d/%m/%Y %I:%M %p")
	
	def draw_footer(self):
		"""Dibujar el pie de página con tres secciones."""
		
		#-- Calcular dimensiones.
		page_width, page_height = self._pagesize
		available_width = page_width - self.left_margin - self.right_margin
		
		#-- Margen inferior.
		y_position = 20
		
		#-- Establecer la fuente, tamaño y color del pie de página.
		self.footer_font = ("Helvetica", 8)
		self.footer_color = colors.black
		
		#-- Estilos de fuente.
		self.setFont("Helvetica", 8)
		self.setFillColor(colors.black)
		
		#-- Extremo izquierdo: texto fijo.
		self.setFont("Helvetica-Oblique", 8)  # Fuente itálica
		self.drawString(self.left_margin, y_position, "M.A.A.Soft")
		self.setFont("Helvetica", 8)
		
		#-- Centro: numeración de páginas.
		page_text = f"- {self._pageNumber} -"
		self.drawCentredString(self.left_margin + available_width/2, y_position, page_text)
		
		#-- Extremo derecho: fecha del reporte.
		self.drawRightString(self.left_margin + available_width, y_position, self.report_date)
		
		#-- Línea horizontal por encima del pie de página.
		self.line(self.left_margin, y_position + 10, self.left_margin + available_width, y_position + 10)
	
	def showPage(self):
		"""Llama a la función que dibuja elementos comunes en cada página."""
		
		#-- Dibujar el pie de página antes de pasar a la siguiente página.
		self.draw_footer()
		
		#-- Llamar al método showPage original para cambiar de página.
		super().showPage()


#---------------------------------------------------------------------------------
class PDFGenerator:
	def __init__(self, context, pagesize=portrait(A4), margins=(10, 10, 0, 40), body_font_size=6, header_font_size=9):
		#-- Dimensiones formato A4:
		#-- portrait:	595.27x841.88 pts. (210x297mm) vertical
		#-- landscape:	841.88x595.27 pts. (297x210mm) horizontal
		
		self.buffer = BytesIO()
		self.context = context
		self.pagesize = pagesize
		self.left_margin, self.right_margin, self.top_margin, self.bottom_margin = margins
		self.body_font_size = body_font_size
		self.header_font_size = header_font_size
		self.extra_header_height = 0
		
		#-- Creación y Configuración del documento base.
		self.doc = BaseDocTemplate(
			self.buffer,
			pagesize=self.pagesize,
			leftMargin=self.left_margin,
			rightMargin=self.right_margin,
			topMargin=self.top_margin,
			bottomMargin=self.bottom_margin
		)
		
		#-- Asignar el contexto del reporte al documento.
		self.doc.contexto_reporte = self.context
		
		#-- Estilos.
		self.styles = getSampleStyleSheet()
		self._setup_custom_styles()
	
		#-- Creación y Configuración inicial del frame.
		self._setup_frame()
		
	def _setup_frame(self, extra_header_height=0):
		"""Configura el frame con márgenes adaptables"""
		
		HEADER_TOP_HEIGHT = 50
		FOOTER_HEIGHT = 40
		SEPARATION = 5
		
		#-- Calcular espacio total ocupado por headers.
		total_header_height = HEADER_TOP_HEIGHT + extra_header_height + SEPARATION
		
		#-- Calcular altura disponible para el cuerpo.
		usable_height = self.doc.height - total_header_height - FOOTER_HEIGHT
		
		#-- Crear frame con altura ajustada.
		# Posicionamiento preciso del Frame
		self.frame = Frame(
			self.doc.leftMargin,	# X: desde margen izquierdo
			FOOTER_HEIGHT,			# Y: desde el borde inferior (justo arriba del footer)
			self.doc.width,			# Ancho completo disponible
			usable_height,			# Altura calculada
			id="body",
			topPadding=0,
			bottomPadding=0,
			leftPadding=0,
			rightPadding=0
		)
		
		#-- Limpiar plantillas existentes (evita duplicados).
		if hasattr(self.doc, 'pageTemplates'):
			self.doc.pageTemplates.clear()  # Elimina plantillas viejas
		
		#-- Crear la plantilla nueva.
		self.doc.addPageTemplates([
			PageTemplate(
				id="reportTemplate",
				frames=[self.frame],
				onPage=self._header_footer
			)
		])
		
	def _setup_custom_styles(self):
		#-- Estilos personalizados.
		
		#-- Estilo de párrafo para los datos en la tabla del reporte.
		self.styles.add(ParagraphStyle(
			name='CellStyle',
			parent=self.styles["BodyText"],
			fontSize=self.body_font_size,
			leading=8,
			spaceBefore=0,
			spaceAfter=0,
			leftIndent=0,
			rightIndent=0,
			firstLineIndent=0,
		))
		
		#-- Estilo de párrafo para la sección Header-bottom-left.
		self.styles.add(ParagraphStyle(
			name='HeaderBottomLeft',
			fontSize=self.header_font_size,
			leading=12,
			alignment=TA_LEFT
		))
		
		#-- Estilo de párrafo para la sección Header-bottom-right.
		self.styles.add(ParagraphStyle(
			name='HeaderBottomRight',
			fontSize=self.header_font_size,
			leading=12,
			alignment=TA_RIGHT
		))
	
	def _get_header_bottom_left(self, context):
		"""SOBREESCRIBIR ESTE MÉTODO PARA CONTENIDO IZQUIERDO PERSONALIZADO"""
		return context.get("header_bottom_left", "")
	
	def _get_header_bottom_right(self, context):
		"""SOBREESCRIBIR ESTE MÉTODO PARA CONTENIDO DERECHO PERSONALIZADO"""
		params = context.get("parametros", {})
		return "<br/>".join([f"<b>{k}:</b> {v}" for k, v in params.items()])
	
	def _calculate_header_bottom_height(self):
		"""Calcula la altura necesaria para el header-bottom"""
		
		#-- Medir contenido izquierdo.
		left_content = self._get_header_bottom_left(self.context)
		left_para = Paragraph(left_content, self.styles['HeaderBottomLeft'])
		
		#-- Medir contenido derecho.
		right_content = self._get_header_bottom_right(self.context)
		right_para = Paragraph(right_content, self.styles['HeaderBottomRight'])
		
		#-- Calcular dimensiones (usando un canvas temporal).
		#-- doc.width ya incluye el descuento de márgenes (ver __init__).
		temp_canvas = Canvas(self.buffer)
		available_width = self.doc.width / 2.0
		
		left_para.wrapOn(temp_canvas, available_width, self.doc.height)
		right_para.wrapOn(temp_canvas, available_width, self.doc.height)
		
		#-- Calcular altura extra necesaria (30pt es la altura base).
		extra_height = max(left_para.height, right_para.height) - 30
		
		return max(0, extra_height)  #-- No permitir valores negativos.
	
	# -------------------------------------------------------------------------------------
	
	def _header_footer(self, canvas_obj, doc):
		"""Método para dibujar header y footer en cada página"""
		
		canvas_obj.saveState()
		width_total, height_total = doc.pagesize
		
		#-- Header Top (logo + título) (altura fija).
		self._render_header_top(canvas_obj, doc, width_total, height_total)
		
		#-- Header Bottom (altura dinámica).
		self._render_header_bottom(canvas_obj, doc, width_total, height_total)
		
		#-- Footer (altura fija).
		self._render_footer(canvas_obj, doc, width_total)
		
		canvas_obj.restoreState()
		
	def _render_header_top(self, canvas_obj, doc, width_total, height_total):
		# --- Header-top -------------------------------------------------------------------
		
		#-- Altura fija de la sección header-top.
		header_top_height = 50
		
		#-- Sección Superior Izquierda: logotipo.
		
		#-- Obtener la URL del logo del contexto.
		logo_path = doc.contexto_reporte.get("logo_path", "")
		
		#-- Coordenadas para renderizar el logo.
		logo_height = 30
		logo_width = 100
		logo_x = doc.leftMargin
		logo_y = height_total - header_top_height + (header_top_height - logo_height) / 2
		
		#-- Usar la ruta del logo si existe.
		if logo_path and os.path.exists(logo_path):
			try:
				canvas_obj.drawImage(
					logo_path, 
					logo_x, 
					logo_y, 
					width=logo_width, 
					height=logo_height, 
					preserveAspectRatio=True, 
					mask='auto'
				)
			except Exception:
				#-- Si falla, no mostrar nada
				pass
		
		#-- Sección Superior Perecha: título.
		titulo = doc.contexto_reporte.get("titulo", "Reporte")
		canvas_obj.setFont("Helvetica-BoldOblique", 12)
		canvas_obj.drawRightString(
			width_total - doc.rightMargin-10, 
			(height_total - header_top_height/2)-10, 
			titulo
		)
	
	def _render_header_bottom(self, canvas_obj, doc, width, height):
		"""Renderiza el header-bottom en posición fija respecto al header-top"""
		
		#-- Posición FIJA: 50pt desde el borde superior (header-top) + margen.
		line_y_start = height - 50
		
		#-- Dibujar línea superior del header-bottom.
		canvas_obj.setLineWidth(1)
		canvas_obj.line(doc.leftMargin, line_y_start, width - doc.rightMargin, line_y_start)
		
		#-- Renderizar contenido (posición relativa al borde superior).
		content_top = line_y_start - 2  #-- 2pt de margen bajo la línea.
		
		#-- Renderizar contenido (crece hacia abajo).
		self._draw_header_bottom_content(
			canvas_obj, doc,
			self._get_header_bottom_left(doc.contexto_reporte),
			self._get_header_bottom_right(doc.contexto_reporte),
			content_top
		)
	
	def _render_footer(self, canvas_obj, doc, width):
		# --- Footer -----------------------------------------------------------------------
		
		footer_y = 15
		
		#-- Línea decorativa.
		line_y = footer_y + 12
		canvas_obj.setLineWidth(1)
		canvas_obj.setStrokeColor(colors.black)
		canvas_obj.line(doc.leftMargin, line_y, width - doc.rightMargin, line_y)
		
		#-- Texto a la izquierda del footer.
		canvas_obj.setFont("Helvetica-Oblique", 9)
		canvas_obj.drawString(doc.leftMargin, footer_y, "M.A.A.Soft")
		
		#-- Número de página formateado al centro.
		numero_pagina_text = f"Página {canvas_obj._pageNumber}"
		canvas_obj.setFont("Helvetica", 9)
		canvas_obj.drawCentredString(width/2.0, footer_y, numero_pagina_text)
		
		#-- Fecha y hora del reporte a la derecha.
		fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M")
		canvas_obj.setFont("Helvetica", 9)
		canvas_obj.drawRightString(width - doc.rightMargin, footer_y, fecha_reporte)
	
	def _draw_header_bottom_content(self, canvas_obj, doc, left_content, right_content, start_y):
		"""Dibuja el contenido del header-bottom desde la posición start_y hacia abajo"""
		
		p_left = Paragraph(left_content, self.styles['HeaderBottomLeft'])
		p_right = Paragraph(right_content, self.styles['HeaderBottomRight'])
		
		available_width = doc.width / 2.0
		
		#-- Medir contenido.
		p_left.wrapOn(canvas_obj, available_width, doc.height)
		p_right.wrapOn(canvas_obj, available_width, doc.height)
		
		#-- Dibujar contenido alineado desde start_y hacia abajo.
		content_bottom = start_y - max(p_left.height, p_right.height)
		p_left.drawOn(canvas_obj, doc.leftMargin + 10, content_bottom)
		p_right.drawOn(canvas_obj, doc.leftMargin + available_width - 10, content_bottom)
		
		#-- Línea inferior con 2pt de separación.
		line_y = content_bottom - 2
		canvas_obj.line(doc.leftMargin, line_y, doc.width + doc.rightMargin, line_y)
	
	def _create_table(self, data, col_widths, style_config, repeat_rows=1):
		"""Crea una tabla con los datos y estilos proporcionados"""
		table = LongTable(data, colWidths=col_widths, repeatRows=repeat_rows)
		
		#-- Estilo base.
		table_style = TableStyle([
			#-- Estilos comunes toda la tabla.
			('VALIGN', (0,0), (-1,-1), 'TOP'),
			('FONTSIZE', (0,0), (-1,-1), self.body_font_size),
			('LEADING', (0,0), (-1,-1), 8),
			# ('RIGHTPADDING', (0,0), (-1,-1), 2),  # valor por defecto es 6
			# ('LEFTPADDING', (0,0), (-1,-1), 2),  # valor por defecto es 6
			
			#-- Padding de la tabla exceptuando la primera fila (headers).
			('TOPPADDING', (0,repeat_rows), (-1,-1), 1),
			('BOTTOMPADDING', (0,repeat_rows), (-1,-1), 1),
			
			#-- Estilos para la primera fila (headers).
			# ('BACKGROUND', (0,0), (-1,(repeat_rows-1)), colors.gray),
			('BACKGROUND', (0,0), (-1,(repeat_rows-1)), colors.lightgrey),
			('FONTNAME', (0,0), (-1,(repeat_rows-1)), 'Helvetica-Bold'),
			# ('TEXTCOLOR', (0,0), (-1,(repeat_rows-1)), colors.white),
			('TEXTCOLOR', (0,0), (-1,(repeat_rows-1)), colors.black),
			('TOPPADDING', (0,0), (-1,(repeat_rows-1)), 3),
			('BOTTOMPADDING', (0,0), (-1,(repeat_rows-1)), 3),
		])
		
		#-- Añadir configuraciones de estilo adicionales.
		for config in style_config:
			table_style.add(*config)
		
		table.setStyle(table_style)
		return table
	
	def generate(self, table_data, col_widths, table_style_config, repeat_rows=1):
		"""Genera el PDF con los datos proporcionados"""
		
		#-- Pre-calcular altura del header ANTES de construir.
		self.extra_header_height = self._calculate_header_bottom_height()		
		
		#-- Reconfigurar frame con la altura calculada.
		self._setup_frame(self.extra_header_height)
		
		#-- Generar contenido.
		content = []
		
		#-- Crear tabla.
		table = self._create_table(table_data, col_widths, table_style_config, repeat_rows)
		content.append(table)
		
		#-- Construir documento.
		self.doc.build(content)
		
		pdf = self.buffer.getvalue()
		self.buffer.close()
		
		return pdf


def add_row_table(table_data, objetos, fields, table_info, generator):
	"""
	Agrega filas a la lista table_data basándose en los objetos y campos proporcionados, formateando los valores según table_info para la exportación a PDF.
	(Modelos de los CRUDs).
	
	Parámetros:
		table_data (list): La lista a la cual se agregarán las nuevas filas.
		objetos (iterable): Los objetos de datos (usualmente diccionarios) de donde extraer los valores de las filas.
		fields (list): La lista de nombres de campos a extraer de cada objeto.
		table_info (dict): Configuración para cada campo, incluyendo opciones de formato.
		generator: Instancia del generador PDF que proporciona estilos para los Párrafos.
	"""
	
	for obj in objetos:
		row = []
		
		for field in fields:
			value = obj[field]
			field_config = table_info.get(field)
			
			if field_config.get('pdf_paragraph'):
				row.append(Paragraph(str(value) if value else "", generator.styles['CellStyle']))
			else:
				if field_config.get('date_format'):
					row.append(format_date(value) if value else "")
					
				elif isinstance(obj[field], bool):
					if 'estatus' in field:
						row.append("Activo" if value else "Inactivo")
					else:
						row.append("Si" if value else "No")
						
				elif field_config.get('type') == "string": 
					row.append(str(value))
					
				elif field_config.get('type') == "int": 
					row.append(formato_argentino_entero(value))
					
				elif field_config.get('type') in ("float", "decimal"):
					row.append(formato_argentino(value))
					
				else:
					row.append(value if value else "")
		
		table_data.append(row)


class JerarquiaSucursal:
	
	@staticmethod
	def puede_consultar_caja(user, caja):
		"""
		Verifica si un usuario puede consultar una caja específica.
		
		Args:
			user: Usuario autenticado
			caja: Objeto Caja a consultar
			
		Returns:
			bool: True si tiene permiso, False si no
		"""
		
		if not user or not caja:
			return False
		
		jerarquia_usuario = getattr(user, 'jerarquia', 'Z')
		
		#-- Jerarquía 'A' tiene acceso total.
		if jerarquia_usuario in JERARQUIAS_CON_ACCESO_TOTAL:
			return True
		
		#-- Para jerarquías menores, verificar sucursal.
		sucursal_usuario = getattr(user, 'id_sucursal', None)
		sucursal_caja = getattr(caja, 'id_sucursal', None)
		
		#-- Si el usuario no tiene sucursal asignada, no puede consultar.
		if not sucursal_usuario:
			return False
		
		#-- Si la caja no tiene sucursal, no se puede consultar.
		if not sucursal_caja:
			return False
		
		#-- Comparar las sucursales.
		return sucursal_usuario.id_sucursal == sucursal_caja.id_sucursal
	
	@staticmethod
	def verificar_permiso_caja(user, caja_obj):
		"""
		Versión que lanza excepción en lugar de retornar booleano.
		"""
		
		if not JerarquiaSucursal.puede_consultar_caja(user, caja_obj):
			raise PermissionDenied(
				"No tiene permisos para consultar esta caja. "
				"Solo puede consultar cajas de su sucursal."
			)		